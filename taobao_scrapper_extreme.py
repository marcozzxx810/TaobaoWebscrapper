import os
import re
import json
import time 
import random

import requests
import pandas as pd
from retrying import retry
import openpyxl

from login import TaoBaoLogin


requests.packages.urllib3.disable_warnings()

req_session = requests.Session()

GOODS_EXCEL_PATH = 'taobao_goods.xlsx'

DATE=time.strftime("%d-%b-%Y", time.gmtime())

class GoodsSpider:

    def __init__(self, q):
        self.q = q
     
        self.timeout = 15
        self.goods_list = []
        
        tbl = TaoBaoLogin(req_session)
        tbl.login()

    @retry(stop_max_attempt_number=3)
    def spider_goods(self, page):
        """

        :param page: taobao page
        :return:
        """
        s = page * 44
        
        search_url = f'https://s.taobao.com/search?initiative_id=tbindexz_20170306&ie=utf8&spm=a21bo.2017.201856-taobao-item.2&sourceId=tb.index&search_type=item&ssid=s5-e&commend=all&imgfile=&q={self.q}&suggest=history_1&_input_charset=utf-8&wq=biyunt&suggest_query=biyunt&source=suggest&bcoffset=4&p4ppushleft=%2C48&s={s}&data-key=s&data-value={s + 44}'

        # not using coz i dont have proper
        proxies = {'http': '118.24.172.149:1080',
                   'https': '60.205.202.3:3128'
                   }
        
        headers = {
            'referer': 'https://www.taobao.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'
        }
        response = req_session.get(search_url, headers=headers,
                                   verify=False, timeout=self.timeout)
        # print(response.text)
        goods_match = re.search(r'g_page_config = (.*?)}};', response.text)
        
        if not goods_match:
            print('提取页面中的数据失败！')
            print(response.text)
            raise RuntimeError
        goods_str = goods_match.group(1) + '}}'
        goods_list = self._get_goods_info(goods_str)
        self._save_excel(goods_list)
        

    def _get_goods_info(self, goods_str):
        """
        analysis json file and grap the data we needed
        :param goods_str: string
        :return:
        """
        goods_json = json.loads(goods_str)
        goods_items = goods_json['mods']['itemlist']['data']['auctions']
        goods_list = []
        for goods_item in goods_items:
            goods = {'title': goods_item['raw_title'],
                     'price': goods_item['view_price'],
                     'location': goods_item['item_loc'],
                     'sales': goods_item['view_sales'],
                     'comment_url': goods_item['comment_url']}
            goods_list.append(goods)
        return goods_list

    def _save_excel(self, goods_list):
        """
        save json to excel
        :param goods_list: goods data
        :param startrow: 
        :return:
        """
        
        if os.path.exists(GOODS_EXCEL_PATH):
            df = pd.read_excel(GOODS_EXCEL_PATH)
            df = df.append(goods_list)
        else:
            df = pd.DataFrame(goods_list)
        if not os.path.exists(GOODS_EXCEL_PATH):
            writer = pd.ExcelWriter(GOODS_EXCEL_PATH)
            
            df.to_excel(excel_writer=writer, columns=['title', 'price', 'location', 'sales', 'comment_url'], index=False,
                        encoding='utf-8', sheet_name=DATE+self.q)
            writer.save()
            writer.close()
        else :
            excel = openpyxl.load_workbook(GOODS_EXCEL_PATH)
            writer = pd.ExcelWriter(GOODS_EXCEL_PATH, engine='openpyxl')
            writer.book = excel
            writer.sheets = dict((ws.title, ws) for ws in excel.worksheets)

            df.to_excel(writer, columns=['title', 'price', 'location', 'sales', 'comment_url'], index=False,
                        encoding='utf-8', sheet_name=DATE+self.q)

            writer.save()
            writer.close()


    def patch_spider_goods(self):
        """
        patch goods slowly
        
        :return:
        """
        print(DATE+self.q)
       
        if os.path.exists(GOODS_EXCEL_PATH):
            excel_file = openpyxl.load_workbook(GOODS_EXCEL_PATH)
            # print(excel_file.sheetnames()) 
            if (DATE+self.q) in excel_file.sheetnames:
                re_write_flage = str(input("TYPE LETTER　T to rewrite, TYPE LETTER　F to cancel: "))
                if re_write_flage is 'F':
                    print("Script is terminated")
                    excel_file.close()
                    os._exit(0)
                elif re_write_flage is 'T':
                    print("Script is going to rewrite")
                    excel_file.remove(excel_file[DATE+self.q])
                    excel_file.create_sheet(DATE+self.q)
                excel_file.save(GOODS_EXCEL_PATH)
                excel_file.close()
        
        for i in range(0, 2):
            print('第%d页' % (i + 1))
            self.spider_goods(i)
            
            time.sleep(random.randint(10, 15))


if __name__ == '__main__':
    gs = GoodsSpider('Xiaomi10')
    gs.patch_spider_goods()

    time.sleep(random.randint(30, 60))
    gs = GoodsSpider('Xiaomi10Pro')
    gs.patch_spider_goods()
